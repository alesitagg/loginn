from flask import Flask, request, redirect
import openpyxl
import os

app = Flask(__name__)
EXCEL_FILE = 'users.xlsx'

# Crear archivo si no existe
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Usuario", "Contraseña", "Nombre", "Puesto"])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    with open('index.html', encoding='utf-8') as f:
        return f.read()

@app.route('/register', methods=['POST'])
def register():
    username = request.form['username']
    password = request.form['password']
    nombre = request.form['nombre']
    puesto = request.form['puesto']

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == username:
            return "Usuario ya registrado"

    sheet.append([username, password, nombre, puesto])
    wb.save(EXCEL_FILE)
    return redirect('/')

@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == username and row[1].value == password:
            return f"Bienvenido, {row[2].value} ({row[3].value})"
    return "Usuario o contraseña incorrecta"

if __name__ == '__main__':
    app.run(debug=True)
